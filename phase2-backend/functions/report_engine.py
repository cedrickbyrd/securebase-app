"""
Report Engine Lambda - Phase 4
Handles analytics queries, report generation, and export functionality
"""

import json
import boto3
import os
from datetime import datetime, timedelta
from decimal import Decimal
import logging
from typing import Dict, List, Any, Optional
import csv
import io
import base64
from io import BytesIO

# Setup logging
logger = logging.getLogger()
logger.setLevel(os.environ.get('LOG_LEVEL', 'INFO'))

# AWS clients
dynamodb = boto3.resource('dynamodb')
s3 = boto3.client('s3')

# Environment variables with validation
REPORTS_TABLE = os.environ.get('REPORTS_TABLE', 'securebase-dev-reports')
SCHEDULES_TABLE = os.environ.get('SCHEDULES_TABLE', 'securebase-dev-report-schedules')
METRICS_TABLE = os.environ.get('METRICS_TABLE', 'securebase-dev-metrics')
CACHE_TABLE = os.environ.get('CACHE_TABLE', 'securebase-dev-report-cache')
S3_BUCKET = os.environ.get('S3_BUCKET', 'securebase-dev-reports')
ENVIRONMENT = os.environ.get('ENVIRONMENT', 'dev')

# Validate required environment variables
required_vars = ['REPORTS_TABLE', 'METRICS_TABLE', 'CACHE_TABLE', 'S3_BUCKET']
missing_vars = [var for var in required_vars if not os.environ.get(var)]
if missing_vars:
    logger.warning(f"Missing environment variables: {missing_vars}. Using defaults.")

# DynamoDB tables
reports_table = dynamodb.Table(REPORTS_TABLE)
schedules_table = dynamodb.Table(SCHEDULES_TABLE)
metrics_table = dynamodb.Table(METRICS_TABLE)
cache_table = dynamodb.Table(CACHE_TABLE)


class DecimalEncoder(json.JSONEncoder):
    """Helper to convert Decimal to float/int for JSON serialization"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj) if obj % 1 else int(obj)
        return super(DecimalEncoder, self).default(obj)


def lambda_handler(event, context):
    """Main Lambda handler"""
    try:
        logger.info(f"Event: {json.dumps(event)}")
        
        # Parse request
        http_method = event.get('httpMethod', 'GET')
        path = event.get('path', '')
        body = json.loads(event.get('body', '{}')) if event.get('body') else {}
        query_params = event.get('queryStringParameters', {}) or {}
        path_params = event.get('pathParameters', {}) or {}
        
        # Get customer ID from authorizer context
        customer_id = event.get('requestContext', {}).get('authorizer', {}).get('customerId')
        if not customer_id:
            return error_response('Unauthorized', 401)
        
        # Route to appropriate handler
        if 'GET' == http_method:
            if '/analytics' in path:
                if '/summary' in path:
                    return get_summary(customer_id, query_params)
                elif '/cost-breakdown' in path:
                    return get_cost_breakdown(customer_id, query_params)
                elif '/security' in path:
                    return get_security_analytics(customer_id, query_params)
                elif '/compliance' in path:
                    return get_compliance_analytics(customer_id, query_params)
                elif '/usage' in path:
                    return get_usage_analytics(customer_id, query_params)
                else:
                    return get_analytics(customer_id, query_params)
            elif '/reports' in path:
                if path_params.get('reportId'):
                    return get_report(customer_id, path_params['reportId'])
                else:
                    return list_reports(customer_id, query_params)
        
        elif 'POST' == http_method:
            if '/analytics/export' in path:
                return export_report(customer_id, body)
            elif '/reports/schedule' in path:
                return schedule_report(customer_id, body)
            elif '/reports/templates' in path and path_params.get('templateId'):
                return create_from_template(customer_id, path_params['templateId'], body)
            elif '/reports' in path:
                return create_report(customer_id, body)
        
        elif 'PUT' == http_method:
            if path_params.get('reportId'):
                return update_report(customer_id, path_params['reportId'], body)
        
        elif 'DELETE' == http_method:
            if path_params.get('reportId'):
                return delete_report(customer_id, path_params['reportId'])
        
        return error_response('Not Found', 404)
        
    except Exception as e:
        logger.error(f"Error: {str(e)}", exc_info=True)
        return error_response(f'Internal server error: {str(e)}', 500)


def get_analytics(customer_id: str, params: Dict) -> Dict:
    """Get analytics data with filtering and grouping"""
    try:
        date_range = params.get('dateRange', '30d')
        dimension = params.get('dimension', 'service')
        
        # Check cache first
        cache_key = f"{customer_id}#{date_range}#{dimension}"
        cached_result = get_from_cache(cache_key)
        if cached_result:
            logger.info("Returning cached result")
            return success_response(cached_result)
        
        # Calculate date range
        start_date, end_date = parse_date_range(date_range)
        
        # Query metrics table
        metrics = query_metrics(customer_id, start_date, end_date)
        
        # Aggregate by dimension
        aggregated = aggregate_by_dimension(metrics, dimension)
        
        # Calculate summary stats
        summary = calculate_summary(metrics)
        
        result = {
            'dateRange': date_range,
            'dimension': dimension,
            'startDate': start_date.isoformat(),
            'endDate': end_date.isoformat(),
            'summary': summary,
            'data': aggregated,
            'metadata': {
                'recordCount': len(metrics),
                'aggregatedCount': len(aggregated),
                'generatedAt': datetime.utcnow().isoformat(),
            }
        }
        
        # Cache result for 1 hour
        cache_result(cache_key, result, ttl=3600)
        
        return success_response(result)
        
    except Exception as e:
        logger.error(f"Get analytics error: {str(e)}")
        raise


def get_summary(customer_id: str, params: Dict) -> Dict:
    """Get summary metrics"""
    try:
        date_range = params.get('dateRange', '30d')
        start_date, end_date = parse_date_range(date_range)
        
        # Query metrics
        metrics = query_metrics(customer_id, start_date, end_date)
        
        # Calculate summary
        summary = calculate_summary(metrics)
        
        return success_response(summary)
        
    except Exception as e:
        logger.error(f"Get summary error: {str(e)}")
        raise


def get_cost_breakdown(customer_id: str, params: Dict) -> Dict:
    """Get cost breakdown by dimension"""
    try:
        date_range = params.get('dateRange', '30d')
        dimension = params.get('dimension', 'service')
        
        start_date, end_date = parse_date_range(date_range)
        metrics = query_metrics(customer_id, start_date, end_date)
        
        # Group by dimension and calculate costs
        breakdown = {}
        for metric in metrics:
            key = metric.get(dimension, 'unknown')
            cost = float(metric.get('cost', 0))
            
            if key not in breakdown:
                breakdown[key] = {'cost': 0, 'count': 0}
            
            breakdown[key]['cost'] += cost
            breakdown[key]['count'] += 1
        
        # Convert to list and sort by cost
        breakdown_list = [
            {'dimension': k, 'cost': v['cost'], 'count': v['count']}
            for k, v in breakdown.items()
        ]
        breakdown_list.sort(key=lambda x: x['cost'], reverse=True)
        
        total_cost = sum(item['cost'] for item in breakdown_list)
        
        return success_response({
            'breakdown': breakdown_list,
            'totalCost': total_cost,
            'dimension': dimension,
            'dateRange': date_range,
        })
        
    except Exception as e:
        logger.error(f"Get cost breakdown error: {str(e)}")
        raise


def get_security_analytics(customer_id: str, params: Dict) -> Dict:
    """Get security analytics (placeholder for Phase 4)"""
    # TODO: Integrate with GuardDuty, Security Hub
    return success_response({
        'findings': [],
        'severity': {'critical': 0, 'high': 0, 'medium': 0, 'low': 0},
        'trends': [],
    })


def get_compliance_analytics(customer_id: str, params: Dict) -> Dict:
    """Get compliance analytics (placeholder for Phase 4)"""
    # TODO: Integrate with AWS Config
    return success_response({
        'frameworks': [],
        'score': 0,
        'controls': {'passing': 0, 'failing': 0},
    })


def get_usage_analytics(customer_id: str, params: Dict) -> Dict:
    """Get usage analytics (placeholder for Phase 4)"""
    # TODO: Query usage metrics
    return success_response({
        'apiCalls': 0,
        'dataTransfer': 0,
        'activeResources': 0,
    })


def list_reports(customer_id: str, params: Dict) -> Dict:
    """List saved reports"""
    try:
        response = reports_table.query(
            KeyConditionExpression='customer_id = :cid',
            ExpressionAttributeValues={':cid': customer_id},
            ScanIndexForward=False,  # Most recent first
        )
        
        reports = response.get('Items', [])
        
        return success_response({'reports': reports})
        
    except Exception as e:
        logger.error(f"List reports error: {str(e)}")
        raise


def create_report(customer_id: str, data: Dict) -> Dict:
    """Create a new report"""
    try:
        report_id = f"rpt_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
        
        report = {
            'customer_id': customer_id,
            'id': report_id,
            'name': data.get('name'),
            'dateRange': data.get('dateRange'),
            'dimension': data.get('dimension'),
            'config': data.get('config', {}),
            'createdAt': datetime.utcnow().isoformat(),
            'updatedAt': datetime.utcnow().isoformat(),
        }
        
        reports_table.put_item(Item=report)
        
        return success_response(report, status_code=201)
        
    except Exception as e:
        logger.error(f"Create report error: {str(e)}")
        raise


def export_report(customer_id: str, data: Dict) -> Dict:
    """Export report in specified format (CSV, JSON, PDF, Excel)"""
    try:
        format_type = data.get('format', 'csv').lower()
        report_data = data.get('data', [])
        report_name = data.get('name', 'report')
        
        # Generate filename
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_name}_{timestamp}.{format_type}"
        
        if format_type == 'csv':
            return export_csv(report_data, filename)
        elif format_type == 'json':
            return export_json(report_data, filename)
        elif format_type == 'pdf':
            return export_pdf(report_data, report_name, filename)
        elif format_type == 'excel' or format_type == 'xlsx':
            return export_excel(report_data, filename)
        else:
            return error_response(f'Unsupported format: {format_type}', 400)
    
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        return error_response(f'Export failed: {str(e)}', 500)


def export_csv(data: List[Dict], filename: str) -> Dict:
    """Export data as CSV"""
    try:
        if not data:
            return error_response('No data to export', 400)
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        
        # Convert Decimal values to float for CSV
        for row in data:
            clean_row = {}
            for key, value in row.items():
                if isinstance(value, Decimal):
                    clean_row[key] = float(value)
                else:
                    clean_row[key] = value
            writer.writerow(clean_row)
        
        csv_content = output.getvalue()
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'text/csv',
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*',
            },
            'body': csv_content,
            'isBase64Encoded': False,
        }
    
    except Exception as e:
        logger.error(f"CSV export error: {str(e)}")
        raise


def export_json(data: List[Dict], filename: str) -> Dict:
    """Export data as JSON"""
    try:
        json_content = json.dumps(data, cls=DecimalEncoder, indent=2)
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*',
            },
            'body': json_content,
            'isBase64Encoded': False,
        }
    
    except Exception as e:
        logger.error(f"JSON export error: {str(e)}")
        raise


def export_pdf(data: List[Dict], report_name: str, filename: str) -> Dict:
    """Export data as PDF using ReportLab"""
    try:
        # Import ReportLab (requires layer or package in Lambda)
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            # Fallback: Return HTML-based PDF alternative
            logger.warning("ReportLab not available, returning HTML format")
            return export_pdf_html(data, report_name, filename)
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        title = Paragraph(report_name, title_style)
        elements.append(title)
        
        # Timestamp
        timestamp = Paragraph(
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
            styles['Normal']
        )
        elements.append(timestamp)
        elements.append(Spacer(1, 12))
        
        # Data table
        if data:
            # Prepare table data
            headers = list(data[0].keys())
            table_data = [headers]
            
            for row in data[:50]:  # Limit to 50 rows for PDF
                row_data = []
                for key in headers:
                    value = row.get(key, '')
                    if isinstance(value, Decimal):
                        value = float(value)
                    row_data.append(str(value))
                table_data.append(row_data)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        # Encode as base64 for API Gateway
        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/pdf',
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*',
            },
            'body': pdf_base64,
            'isBase64Encoded': True,
        }
    
    except Exception as e:
        logger.error(f"PDF export error: {str(e)}")
        raise


def export_pdf_html(data: List[Dict], report_name: str, filename: str) -> Dict:
    """Fallback PDF export using HTML (client-side rendering)"""
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{report_name}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            h1 {{ color: #1f2937; text-align: center; }}
            .timestamp {{ text-align: center; color: #6b7280; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background-color: #3b82f6; color: white; padding: 12px; text-align: left; }}
            td {{ border: 1px solid #ddd; padding: 8px; }}
            tr:nth-child(even) {{ background-color: #f9fafb; }}
        </style>
    </head>
    <body>
        <h1>{report_name}</h1>
        <div class="timestamp">Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}</div>
        <table>
            <thead>
                <tr>
    """
    
    if data:
        # Add headers
        for key in data[0].keys():
            html_content += f"<th>{key}</th>"
        html_content += "</tr></thead><tbody>"
        
        # Add rows
        for row in data[:100]:  # Limit to 100 rows
            html_content += "<tr>"
            for value in row.values():
                if isinstance(value, Decimal):
                    value = float(value)
                html_content += f"<td>{value}</td>"
            html_content += "</tr>"
    
    html_content += """
            </tbody>
        </table>
        <script>
            // Auto-print on load for PDF generation
            window.onload = function() { window.print(); }
        </script>
    </body>
    </html>
    """
    
    return {
        'statusCode': 200,
        'headers': {
            'Content-Type': 'text/html',
            'Content-Disposition': f'inline; filename="{filename.replace(".pdf", ".html")}"',
            'Access-Control-Allow-Origin': '*',
        },
        'body': html_content,
        'isBase64Encoded': False,
    }


def export_excel(data: List[Dict], filename: str) -> Dict:
    """Export data as Excel using openpyxl"""
    try:
        # Import openpyxl (requires layer or package in Lambda)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # Fallback to CSV with .xlsx extension
            logger.warning("openpyxl not available, falling back to CSV")
            return export_csv(data, filename.replace('.xlsx', '.csv'))
        
        if not data:
            return error_response('No data to export', 400)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Add headers
        headers = list(data[0].keys())
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        for row in data:
            row_data = []
            for key in headers:
                value = row.get(key, '')
                if isinstance(value, Decimal):
                    value = float(value)
                row_data.append(value)
            ws.append(row_data)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        # Encode as base64
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*',
            },
            'body': excel_base64,
            'isBase64Encoded': True,
        }
    
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        raise


def schedule_report(customer_id: str, data: Dict) -> Dict:
    """Schedule a report for automatic delivery"""
    # TODO: Implement with EventBridge/CloudWatch Events
    return success_response({
        'scheduleId': f"sched_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
        'status': 'scheduled',
    })


# Helper functions

def parse_date_range(date_range: str) -> tuple:
    """Parse date range string into start/end dates"""
    end_date = datetime.utcnow()
    
    if date_range == '7d':
        start_date = end_date - timedelta(days=7)
    elif date_range == '30d':
        start_date = end_date - timedelta(days=30)
    elif date_range == '90d':
        start_date = end_date - timedelta(days=90)
    elif date_range == '12m':
        start_date = end_date - timedelta(days=365)
    else:
        start_date = end_date - timedelta(days=30)
    
    return start_date, end_date


def query_metrics(customer_id: str, start_date: datetime, end_date: datetime) -> List:
    """Query metrics from DynamoDB"""
    # Mock data for Phase 4 development
    # TODO: Replace with actual DynamoDB query
    return [
        {
            'customer_id': customer_id,
            'timestamp': (start_date + timedelta(days=i)).isoformat(),
            'service': 'lambda' if i % 2 == 0 else 'dynamodb',
            'region': 'us-east-1',
            'cost': Decimal(str(10 + i)),
            'usage': Decimal(str(100 + i * 10)),
        }
        for i in range(30)
    ]


def aggregate_by_dimension(metrics: List, dimension: str) -> List:
    """Aggregate metrics by specified dimension"""
    aggregated = {}
    
    for metric in metrics:
        key = metric.get(dimension, 'unknown')
        
        if key not in aggregated:
            aggregated[key] = {
                'dimension': key,
                'totalCost': 0,
                'totalUsage': 0,
                'count': 0,
            }
        
        aggregated[key]['totalCost'] += float(metric.get('cost', 0))
        aggregated[key]['totalUsage'] += float(metric.get('usage', 0))
        aggregated[key]['count'] += 1
    
    return list(aggregated.values())


def calculate_summary(metrics: List) -> Dict:
    """Calculate summary statistics"""
    total_cost = sum(float(m.get('cost', 0)) for m in metrics)
    total_usage = sum(float(m.get('usage', 0)) for m in metrics)
    
    return {
        'totalCost': total_cost,
        'apiCalls': total_usage,
        'complianceScore': 95,  # Mock
        'activeResources': len(metrics),
        'costChange': 5.2,  # Mock % change
        'apiCallsChange': 12.3,  # Mock
        'complianceChange': 2.1,  # Mock
        'resourcesChange': -1.5,  # Mock
    }


def get_from_cache(cache_key: str) -> Optional[Dict]:
    """Get result from cache"""
    try:
        response = cache_table.get_item(Key={'cache_key': cache_key})
        item = response.get('Item')
        
        if item and item.get('ttl', 0) > int(datetime.utcnow().timestamp()):
            return item.get('data')
        
        return None
    except Exception as e:
        logger.warning(f"Cache get error: {str(e)}")
        return None


def cache_result(cache_key: str, data: Dict, ttl: int = 3600):
    """Cache result with TTL"""
    try:
        cache_table.put_item(Item={
            'cache_key': cache_key,
            'data': data,
            'ttl': int(datetime.utcnow().timestamp()) + ttl,
        })
    except Exception as e:
        logger.warning(f"Cache put error: {str(e)}")


def success_response(data: Any, status_code: int = 200) -> Dict:
    """Return success response"""
    return {
        'statusCode': status_code,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*',
        },
        'body': json.dumps(data, cls=DecimalEncoder),
    }


def error_response(message: str, status_code: int = 400) -> Dict:
    """Return error response"""
    return {
        'statusCode': status_code,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*',
        },
        'body': json.dumps({'error': message}),
    }
