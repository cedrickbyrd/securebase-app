"""
Analytics Reporter Lambda - Phase 4
Generates analytics reports in various formats (CSV, JSON, PDF, Excel)
Handles both on-demand and scheduled report generation
"""

import json
import boto3
import os
from datetime import datetime, timedelta
from decimal import Decimal
import logging
from typing import Dict, List, Any, Optional
import csv
import io
import base64

# Setup logging
logger = logging.getLogger()
logger.setLevel(os.environ.get('LOG_LEVEL', 'INFO'))

# AWS clients
dynamodb = boto3.resource('dynamodb')
s3 = boto3.client('s3')
sns = boto3.client('sns')

# Environment variables
METRICS_TABLE = os.environ.get('METRICS_TABLE', 'securebase-dev-metrics')
REPORTS_TABLE = os.environ.get('REPORTS_TABLE', 'securebase-dev-reports')
S3_BUCKET = os.environ.get('S3_BUCKET', 'securebase-dev-reports')
SNS_TOPIC = os.environ.get('SNS_TOPIC', '')
ENVIRONMENT = os.environ.get('ENVIRONMENT', 'dev')

# DynamoDB tables
metrics_table = dynamodb.Table(METRICS_TABLE)
reports_table = dynamodb.Table(REPORTS_TABLE)


class DecimalEncoder(json.JSONEncoder):
    """Helper to convert Decimal to float/int for JSON serialization"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj) if obj % 1 else int(obj)
        return super(DecimalEncoder, self).default(obj)


def lambda_handler(event, context):
    """
    Main Lambda handler - generates reports based on request or schedule
    Can be triggered by:
    - API Gateway (on-demand reports)
    - EventBridge (scheduled reports)
    """
    try:
        logger.info(f"Event: {json.dumps(event)}")
        
        # Determine trigger source
        if event.get('source') == 'aws.events':
            # Scheduled report from EventBridge
            return handle_scheduled_report(event)
        else:
            # On-demand report from API Gateway
            return handle_api_request(event)
        
    except Exception as e:
        logger.error(f"Report generation failed: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error': str(e)})
        }


def handle_api_request(event: Dict) -> Dict:
    """Handle API Gateway request for on-demand report"""
    try:
        # Parse request
        http_method = event.get('httpMethod', 'POST')
        body = json.loads(event.get('body', '{}')) if event.get('body') else {}
        
        # Get customer ID from authorizer
        customer_id = event.get('requestContext', {}).get('authorizer', {}).get('customerId')
        if not customer_id:
            return error_response('Unauthorized', 401)
        
        # Validate request
        report_type = body.get('type', 'monthly')
        format_type = body.get('format', 'pdf')
        period = body.get('period', '30d')
        
        logger.info(f"Generating {format_type} report for customer {customer_id}, period {period}")
        
        # Generate report
        report_data = generate_report(customer_id, report_type, period)
        
        # Format and store
        output = format_report(report_data, format_type, customer_id, report_type)
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'message': 'Report generated successfully',
                'url': output['url'],
                'filename': output['filename'],
                'size_bytes': output['size']
            }, cls=DecimalEncoder)
        }
        
    except Exception as e:
        logger.error(f"API request failed: {str(e)}")
        return error_response(str(e), 500)


def handle_scheduled_report(event: Dict) -> Dict:
    """Handle EventBridge scheduled report generation"""
    try:
        # Get schedule details from event
        schedule_id = event.get('detail', {}).get('schedule_id')
        customer_id = event.get('detail', {}).get('customer_id')
        
        if not customer_id or not schedule_id:
            raise ValueError("Missing schedule_id or customer_id in event")
        
        logger.info(f"Processing scheduled report {schedule_id} for customer {customer_id}")
        
        # Get schedule configuration
        schedule = get_schedule(customer_id, schedule_id)
        if not schedule:
            raise ValueError(f"Schedule {schedule_id} not found")
        
        # Generate report
        report_data = generate_report(
            customer_id,
            schedule.get('report_type', 'monthly'),
            schedule.get('period', '30d')
        )
        
        # Format report
        output = format_report(
            report_data,
            schedule.get('format', 'pdf'),
            customer_id,
            schedule.get('report_type', 'monthly')
        )
        
        # Send notification
        recipients = schedule.get('recipients', [])
        if recipients:
            send_notification(customer_id, output, recipients)
        
        logger.info(f"Scheduled report completed: {output['filename']}")
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Scheduled report generated',
                'url': output['url']
            })
        }
        
    except Exception as e:
        logger.error(f"Scheduled report failed: {str(e)}")
        raise


def generate_report(customer_id: str, report_type: str, period: str) -> Dict[str, Any]:
    """Generate report data from metrics"""
    try:
        # Parse period (e.g., "30d", "7d", "1m")
        days = parse_period(period)
        end_time = datetime.utcnow()
        start_time = end_time - timedelta(days=days)
        
        logger.info(f"Querying metrics from {start_time} to {end_time}")
        
        # Query metrics from DynamoDB
        response = metrics_table.query(
            KeyConditionExpression='customer_id = :cid AND #ts BETWEEN :start AND :end',
            ExpressionAttributeNames={'#ts': 'timestamp'},
            ExpressionAttributeValues={
                ':cid': customer_id,
                ':start': start_time.isoformat(),
                ':end': end_time.isoformat()
            }
        )
        
        metrics = response.get('Items', [])
        logger.info(f"Found {len(metrics)} metrics")
        
        # Aggregate metrics by type
        aggregated = aggregate_metrics(metrics)
        
        # Build report structure
        report = {
            'customer_id': customer_id,
            'report_type': report_type,
            'period': period,
            'generated_at': datetime.utcnow().isoformat(),
            'start_date': start_time.isoformat(),
            'end_date': end_time.isoformat(),
            'summary': calculate_summary(aggregated),
            'metrics': aggregated,
            'trends': calculate_trends(aggregated, days)
        }
        
        return report
        
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise


def format_report(report_data: Dict, format_type: str, customer_id: str, report_type: str) -> Dict[str, Any]:
    """Format report in requested format and upload to S3"""
    try:
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f"{customer_id}/{report_type}_{timestamp}.{format_type}"
        
        if format_type == 'json':
            content = json.dumps(report_data, cls=DecimalEncoder, indent=2)
            content_type = 'application/json'
            
        elif format_type == 'csv':
            content = generate_csv(report_data)
            content_type = 'text/csv'
            
        elif format_type == 'pdf':
            content = generate_pdf(report_data)
            content_type = 'application/pdf'
            
        elif format_type == 'excel':
            content = generate_excel(report_data)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
        else:
            raise ValueError(f"Unsupported format: {format_type}")
        
        # Upload to S3
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=filename,
            Body=content,
            ContentType=content_type,
            ServerSideEncryption='AES256',
            Metadata={
                'customer_id': customer_id,
                'report_type': report_type,
                'generated_at': datetime.utcnow().isoformat()
            }
        )
        
        # Generate presigned URL (valid for 7 days)
        url = s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': S3_BUCKET, 'Key': filename},
            ExpiresIn=604800  # 7 days
        )
        
        logger.info(f"Report uploaded: s3://{S3_BUCKET}/{filename}")
        
        return {
            'url': url,
            'filename': filename,
            'size': len(content) if isinstance(content, (str, bytes)) else 0,
            'format': format_type
        }
        
    except Exception as e:
        logger.error(f"Error formatting report: {str(e)}")
        raise


def generate_csv(report_data: Dict) -> str:
    """Generate CSV format report"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['Analytics Report'])
    writer.writerow(['Customer ID', report_data['customer_id']])
    writer.writerow(['Period', report_data['period']])
    writer.writerow(['Generated', report_data['generated_at']])
    writer.writerow([])
    
    # Summary
    writer.writerow(['Summary'])
    summary = report_data.get('summary', {})
    for key, value in summary.items():
        writer.writerow([key.replace('_', ' ').title(), value])
    writer.writerow([])
    
    # Metrics
    writer.writerow(['Detailed Metrics'])
    writer.writerow(['Metric', 'Value', 'Unit', 'Service'])
    
    for metric_name, metric_data in report_data.get('metrics', {}).items():
        writer.writerow([
            metric_name.replace('_', ' ').title(),
            metric_data.get('value', 0),
            metric_data.get('unit', ''),
            metric_data.get('service', '')
        ])
    
    return output.getvalue()


def generate_pdf(report_data: Dict) -> bytes:
    """
    Generate PDF format report (requires ReportLab layer)
    
    NOTE: This is a simplified implementation for MVP.
    Production version would use ReportLab for rich PDF formatting.
    """
    # TODO: Implement full PDF generation with ReportLab
    # For now, return text-based content that can be displayed
    # The Lambda layer with ReportLab should be attached for full functionality
    
    try:
        # Try to import ReportLab if available in Lambda layer
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(f"<b>Analytics Report</b>", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Content
        content = f"""
        Customer ID: {report_data['customer_id']}<br/>
        Period: {report_data['period']}<br/>
        Generated: {report_data['generated_at']}<br/><br/>
        
        <b>Summary:</b><br/>
        {json.dumps(report_data.get('summary', {}), indent=2, cls=DecimalEncoder)}
        """
        story.append(Paragraph(content.replace('\n', '<br/>'), styles['Normal']))
        
        doc.build(story)
        return buffer.getvalue()
        
    except ImportError:
        # ReportLab not available - return plain text as fallback
        logger.warning("ReportLab not available, generating text-based PDF")
        text_content = f"""
Analytics Report
================

Customer ID: {report_data['customer_id']}
Period: {report_data['period']}
Generated: {report_data['generated_at']}

Summary:
--------
{json.dumps(report_data.get('summary', {}), indent=2, cls=DecimalEncoder)}

Metrics:
--------
{json.dumps(report_data.get('metrics', {}), indent=2, cls=DecimalEncoder)}

NOTE: Install ReportLab Lambda layer for rich PDF formatting.
"""
        return text_content.encode('utf-8')


def generate_excel(report_data: Dict) -> bytes:
    """
    Generate Excel format report (requires openpyxl layer)
    
    NOTE: This is a simplified implementation for MVP.
    Production version would use openpyxl for rich Excel formatting.
    """
    try:
        # Try to import openpyxl if available in Lambda layer
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"
        
        # Header
        ws['A1'] = "Analytics Report"
        ws['A2'] = f"Customer ID: {report_data['customer_id']}"
        ws['A3'] = f"Period: {report_data['period']}"
        ws['A4'] = f"Generated: {report_data['generated_at']}"
        
        # Summary section
        row = 6
        ws[f'A{row}'] = "Summary"
        row += 1
        for key, value in report_data.get('summary', {}).items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
        
        # Metrics section
        row += 2
        ws[f'A{row}'] = "Detailed Metrics"
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Unit"
        row += 1
        
        for metric_name, metric_data in report_data.get('metrics', {}).items():
            ws[f'A{row}'] = metric_name.replace('_', ' ').title()
            ws[f'B{row}'] = float(metric_data.get('value', 0))
            ws[f'C{row}'] = metric_data.get('unit', '')
            row += 1
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
        
    except ImportError:
        # openpyxl not available - return CSV as fallback
        logger.warning("openpyxl not available, generating CSV instead of Excel")
        return generate_csv(report_data).encode('utf-8')


def aggregate_metrics(metrics: List[Dict]) -> Dict[str, Any]:
    """Aggregate raw metrics by metric name"""
    aggregated = {}
    
    for metric in metrics:
        metric_name = metric.get('metric_name')
        if not metric_name:
            continue
        
        if metric_name not in aggregated:
            aggregated[metric_name] = {
                'value': Decimal('0'),
                'count': 0,
                'unit': metric.get('unit', ''),
                'service': metric.get('service', ''),
                'samples': []
            }
        
        value = metric.get('value', Decimal('0'))
        aggregated[metric_name]['value'] += value
        aggregated[metric_name]['count'] += 1
        aggregated[metric_name]['samples'].append({
            'timestamp': metric.get('timestamp'),
            'value': value
        })
    
    # Calculate averages where appropriate
    for metric_name in aggregated:
        if aggregated[metric_name]['count'] > 0:
            aggregated[metric_name]['average'] = (
                aggregated[metric_name]['value'] / aggregated[metric_name]['count']
            )
    
    return aggregated


def calculate_summary(aggregated: Dict) -> Dict[str, Any]:
    """Calculate summary statistics from aggregated metrics"""
    summary = {}
    
    for metric_name, data in aggregated.items():
        if metric_name == 'api_calls':
            summary['total_api_calls'] = int(data['value'])
        elif metric_name == 'storage_gb':
            summary['total_storage_gb'] = float(data['value'])
        elif metric_name == 'compute_hours':
            summary['total_compute_hours'] = float(data['value'])
        elif metric_name == 'daily_cost':
            summary['total_cost'] = float(data['value'])
        elif metric_name == 'compliance_score':
            summary['avg_compliance_score'] = float(data.get('average', 0))
    
    return summary


def calculate_trends(aggregated: Dict, days: int) -> Dict[str, Any]:
    """Calculate trends over the period"""
    trends = {}
    
    # For each metric, calculate trend (increase/decrease)
    for metric_name, data in aggregated.items():
        samples = data.get('samples', [])
        if len(samples) < 2:
            trends[metric_name] = 'insufficient_data'
            continue
        
        # Compare first half vs second half
        mid_point = len(samples) // 2
        first_half = sum(float(s['value']) for s in samples[:mid_point]) / mid_point
        second_half = sum(float(s['value']) for s in samples[mid_point:]) / (len(samples) - mid_point)
        
        if first_half > 0:
            change_pct = ((second_half - first_half) / first_half) * 100
            trends[metric_name] = f"{'+' if change_pct > 0 else ''}{change_pct:.1f}%"
        else:
            trends[metric_name] = 'new_metric'
    
    return trends


def parse_period(period: str) -> int:
    """Parse period string (e.g., '30d', '7d', '1m') into days"""
    period = period.lower().strip()
    
    if period.endswith('d'):
        return int(period[:-1])
    elif period.endswith('w'):
        return int(period[:-1]) * 7
    elif period.endswith('m'):
        return int(period[:-1]) * 30
    elif period.endswith('y'):
        return int(period[:-1]) * 365
    else:
        return 30  # Default to 30 days


def get_schedule(customer_id: str, schedule_id: str) -> Optional[Dict]:
    """Retrieve schedule configuration from DynamoDB"""
    try:
        # In production, query schedules table
        # For now, return mock data
        return {
            'schedule_id': schedule_id,
            'customer_id': customer_id,
            'report_type': 'monthly',
            'format': 'pdf',
            'period': '30d',
            'recipients': ['admin@example.com']
        }
    except Exception as e:
        logger.error(f"Error getting schedule: {str(e)}")
        return None


def send_notification(customer_id: str, output: Dict, recipients: List[str]):
    """Send email notification with report link"""
    try:
        if not SNS_TOPIC:
            logger.warning("SNS topic not configured, skipping notification")
            return
        
        message = f"""
Your analytics report is ready!

Report: {output['filename']}
Format: {output['format']}
Size: {output['size']} bytes

Download: {output['url']}

This link will expire in 7 days.
"""
        
        sns.publish(
            TopicArn=SNS_TOPIC,
            Subject='Analytics Report Ready',
            Message=message
        )
        
        logger.info(f"Notification sent to {len(recipients)} recipients")
        
    except Exception as e:
        logger.error(f"Error sending notification: {str(e)}")


def error_response(message: str, status_code: int = 400) -> Dict:
    """Generate error response"""
    return {
        'statusCode': status_code,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*'
        },
        'body': json.dumps({'error': message})
    }
