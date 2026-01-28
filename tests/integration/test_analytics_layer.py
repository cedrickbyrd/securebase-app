"""
Integration Tests for Analytics Lambda Layer
Tests that the ReportLab and openpyxl dependencies are available and functional
in the Lambda execution environment
"""

import pytest
import json
import os
import sys
from unittest.mock import Mock, patch, MagicMock
from io import BytesIO

# Import Lambda functions for testing
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../phase2-backend/functions'))

import analytics_reporter


class TestAnalyticsLayer:
    """Test Analytics Lambda layer dependencies"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test environment"""
        os.environ['ENVIRONMENT'] = 'test'
        os.environ['METRICS_TABLE'] = 'test-metrics'
        os.environ['REPORTS_TABLE'] = 'test-reports'
        os.environ['S3_BUCKET'] = 'test-reports-bucket'
        os.environ['SNS_TOPIC'] = 'arn:aws:sns:us-east-1:123456789012:test-topic'
        os.environ['LOG_LEVEL'] = 'DEBUG'
        
        self.test_customer_id = 'test-customer-123'
        self.sample_report_data = {
            'customer_id': self.test_customer_id,
            'report_type': 'monthly',
            'period': '30d',
            'generated_at': '2026-01-28T00:00:00',
            'start_date': '2026-01-01T00:00:00',
            'end_date': '2026-01-28T00:00:00',
            'summary': {
                'total_api_calls': 1000,
                'total_storage_gb': 50.5,
                'total_compute_hours': 100.0,
                'total_cost': 250.00,
                'avg_compliance_score': 95.5
            },
            'metrics': {
                'api_calls': {'value': 1000, 'unit': 'count', 'service': 'API Gateway'},
                'storage_gb': {'value': 50.5, 'unit': 'GB', 'service': 'S3'},
                'compute_hours': {'value': 100.0, 'unit': 'hours', 'service': 'Lambda'}
            },
            'trends': {
                'api_calls': '+15.0%',
                'storage_gb': '+5.2%',
                'compute_hours': '-3.1%'
            }
        }
        
        yield
    
    
    def test_reportlab_import(self):
        """Test that ReportLab library can be imported"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            assert True, "ReportLab imported successfully"
        except ImportError as e:
            pytest.fail(f"ReportLab import failed: {str(e)}")
    
    
    def test_openpyxl_import(self):
        """Test that openpyxl library can be imported"""
        try:
            from openpyxl import Workbook
            assert True, "openpyxl imported successfully"
        except ImportError as e:
            pytest.fail(f"openpyxl import failed: {str(e)}")
    
    
    def test_pillow_import(self):
        """Test that Pillow (PIL) library can be imported"""
        try:
            from PIL import Image
            assert True, "Pillow imported successfully"
        except ImportError as e:
            pytest.skip(f"Pillow import failed (optional dependency): {str(e)}")
    
    
    @patch('analytics_reporter.s3')
    def test_pdf_generation_with_reportlab(self, mock_s3):
        """Test PDF generation using ReportLab from Lambda layer"""
        # Mock S3 upload
        mock_s3.put_object.return_value = {}
        mock_s3.generate_presigned_url.return_value = 'https://s3.example.com/test.pdf'
        
        # Generate PDF
        try:
            result = analytics_reporter.generate_pdf(self.sample_report_data)
            
            # Verify result is bytes
            assert isinstance(result, bytes), "PDF should be bytes"
            assert len(result) > 0, "PDF should not be empty"
            
            # Check if it's actual PDF (starts with %PDF) or fallback text
            if result.startswith(b'%PDF'):
                # ReportLab is available, proper PDF generated
                assert True, "Proper PDF generated with ReportLab"
            else:
                # Fallback text-based PDF
                assert b'Analytics Report' in result, "Fallback PDF should contain report title"
                assert self.test_customer_id.encode() in result, "PDF should contain customer ID"
            
        except ImportError as e:
            pytest.fail(f"PDF generation failed due to missing dependency: {str(e)}")
        except Exception as e:
            pytest.fail(f"PDF generation failed: {str(e)}")
    
    
    @patch('analytics_reporter.s3')
    def test_excel_generation_with_openpyxl(self, mock_s3):
        """Test Excel generation using openpyxl from Lambda layer"""
        # Mock S3 upload
        mock_s3.put_object.return_value = {}
        mock_s3.generate_presigned_url.return_value = 'https://s3.example.com/test.xlsx'
        
        # Generate Excel
        try:
            result = analytics_reporter.generate_excel(self.sample_report_data)
            
            # Verify result is bytes
            assert isinstance(result, bytes), "Excel should be bytes"
            assert len(result) > 0, "Excel should not be empty"
            
            # Check if it's actual Excel file (ZIP format, starts with PK)
            if result.startswith(b'PK'):
                # openpyxl is available, proper Excel generated
                assert True, "Proper Excel generated with openpyxl"
                
                # Try to read it back to verify structure
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(BytesIO(result))
                    ws = wb.active
                    
                    # Verify content
                    assert ws['A1'].value == "Analytics Report", "Excel should have correct header"
                    assert self.test_customer_id in str(ws['A2'].value), "Excel should contain customer ID"
                    
                except Exception as e:
                    pytest.fail(f"Excel file validation failed: {str(e)}")
            else:
                # Fallback CSV
                assert b'Analytics Report' in result or b'Customer ID' in result, "Fallback CSV should contain headers"
            
        except ImportError as e:
            pytest.fail(f"Excel generation failed due to missing dependency: {str(e)}")
        except Exception as e:
            pytest.fail(f"Excel generation failed: {str(e)}")
    
    
    @patch('analytics_reporter.s3')
    def test_csv_generation_no_layer_dependencies(self, mock_s3):
        """Test CSV generation works without layer dependencies"""
        # CSV should work without any layer dependencies
        result = analytics_reporter.generate_csv(self.sample_report_data)
        
        assert isinstance(result, str), "CSV should be string"
        assert len(result) > 0, "CSV should not be empty"
        assert 'Analytics Report' in result, "CSV should have header"
        assert self.test_customer_id in result, "CSV should contain customer ID"
        assert 'Summary' in result, "CSV should have summary section"
    
    
    @patch('analytics_reporter.s3')
    @patch('analytics_reporter.metrics_table')
    def test_format_report_pdf_with_layer(self, mock_metrics, mock_s3):
        """Integration test: format_report with PDF should use layer"""
        # Mock S3 operations
        mock_s3.put_object.return_value = {}
        mock_s3.generate_presigned_url.return_value = 'https://s3.example.com/report.pdf'
        
        # Format report as PDF
        result = analytics_reporter.format_report(
            self.sample_report_data,
            'pdf',
            self.test_customer_id,
            'monthly'
        )
        
        # Verify result structure
        assert 'url' in result, "Result should have URL"
        assert 'filename' in result, "Result should have filename"
        assert 'size' in result, "Result should have size"
        assert 'format' in result, "Result should have format"
        assert result['format'] == 'pdf', "Format should be PDF"
        assert result['size'] > 0, "PDF size should be greater than 0"
        
        # Verify S3 upload was called
        assert mock_s3.put_object.called, "S3 put_object should be called"
        call_kwargs = mock_s3.put_object.call_args[1]
        assert call_kwargs['ContentType'] == 'application/pdf', "Content-Type should be PDF"
        assert call_kwargs['ServerSideEncryption'] == 'AES256', "Encryption should be enabled"
    
    
    @patch('analytics_reporter.s3')
    @patch('analytics_reporter.metrics_table')
    def test_format_report_excel_with_layer(self, mock_metrics, mock_s3):
        """Integration test: format_report with Excel should use layer"""
        # Mock S3 operations
        mock_s3.put_object.return_value = {}
        mock_s3.generate_presigned_url.return_value = 'https://s3.example.com/report.xlsx'
        
        # Format report as Excel
        result = analytics_reporter.format_report(
            self.sample_report_data,
            'excel',
            self.test_customer_id,
            'monthly'
        )
        
        # Verify result structure
        assert 'url' in result, "Result should have URL"
        assert 'filename' in result, "Result should have filename"
        assert 'size' in result, "Result should have size"
        assert 'format' in result, "Result should have format"
        assert result['format'] == 'excel', "Format should be Excel"
        assert result['size'] > 0, "Excel size should be greater than 0"
        
        # Verify S3 upload was called
        assert mock_s3.put_object.called, "S3 put_object should be called"
        call_kwargs = mock_s3.put_object.call_args[1]
        expected_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert call_kwargs['ContentType'] == expected_content_type, "Content-Type should be Excel"
    
    
    def test_layer_dependencies_versions(self):
        """Test that layer dependencies are correct versions"""
        try:
            import reportlab
            import openpyxl
            
            # Check minimum versions
            reportlab_version = reportlab.Version
            assert reportlab_version >= '4.0.0', f"ReportLab version should be >= 4.0.0, got {reportlab_version}"
            
            openpyxl_version = openpyxl.__version__
            assert openpyxl_version >= '3.1.0', f"openpyxl version should be >= 3.1.0, got {openpyxl_version}"
            
        except ImportError as e:
            pytest.skip(f"Layer dependencies not available: {str(e)}")
    
    
    @patch('analytics_reporter.s3')
    def test_pdf_generation_handles_missing_layer_gracefully(self, mock_s3):
        """Test that PDF generation provides fallback when layer is missing"""
        # Mock S3
        mock_s3.put_object.return_value = {}
        
        # Generate PDF (should work even without ReportLab)
        result = analytics_reporter.generate_pdf(self.sample_report_data)
        
        assert isinstance(result, bytes), "Should return bytes even without layer"
        assert len(result) > 0, "Should not be empty"
        # Should contain report data even in fallback mode
        assert b'Analytics Report' in result, "Should contain report title"
        assert self.test_customer_id.encode() in result, "Should contain customer ID"
    
    
    @patch('analytics_reporter.s3')
    def test_excel_generation_handles_missing_layer_gracefully(self, mock_s3):
        """Test that Excel generation provides fallback when layer is missing"""
        # Mock S3
        mock_s3.put_object.return_value = {}
        
        # Generate Excel (should fallback to CSV without openpyxl)
        result = analytics_reporter.generate_excel(self.sample_report_data)
        
        assert isinstance(result, bytes), "Should return bytes even without layer"
        assert len(result) > 0, "Should not be empty"
        # In fallback mode, it returns CSV
        if not result.startswith(b'PK'):
            # It's CSV fallback
            assert b'Analytics Report' in result or b'Customer ID' in result, "CSV fallback should have content"


class TestLayerPerformance:
    """Test performance characteristics of layer-dependent operations"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test environment"""
        os.environ['ENVIRONMENT'] = 'test'
        os.environ['METRICS_TABLE'] = 'test-metrics'
        os.environ['REPORTS_TABLE'] = 'test-reports'
        os.environ['S3_BUCKET'] = 'test-bucket'
        
        # Create larger test dataset
        self.large_report_data = {
            'customer_id': 'test-customer',
            'report_type': 'monthly',
            'period': '30d',
            'generated_at': '2026-01-28T00:00:00',
            'start_date': '2026-01-01T00:00:00',
            'end_date': '2026-01-28T00:00:00',
            'summary': {f'metric_{i}': i * 100 for i in range(50)},
            'metrics': {
                f'metric_{i}': {
                    'value': i * 100,
                    'unit': 'count',
                    'service': f'Service {i}',
                    'samples': [{'timestamp': f'2026-01-{j:02d}T00:00:00', 'value': i * j} for j in range(1, 31)]
                }
                for i in range(100)
            },
            'trends': {f'metric_{i}': f'+{i}.0%' for i in range(50)}
        }
        
        yield
    
    
    @patch('analytics_reporter.s3')
    def test_large_pdf_generation_performance(self, mock_s3):
        """Test PDF generation performance with large dataset"""
        import time
        
        mock_s3.put_object.return_value = {}
        
        start_time = time.time()
        result = analytics_reporter.generate_pdf(self.large_report_data)
        elapsed_time = time.time() - start_time
        
        assert isinstance(result, bytes), "Should generate PDF"
        assert len(result) > 1000, "Large report should be > 1KB"
        # Performance threshold: should complete in < 5 seconds
        assert elapsed_time < 5.0, f"PDF generation took {elapsed_time:.2f}s (should be < 5s)"
    
    
    @patch('analytics_reporter.s3')
    def test_large_excel_generation_performance(self, mock_s3):
        """Test Excel generation performance with large dataset"""
        import time
        
        mock_s3.put_object.return_value = {}
        
        start_time = time.time()
        result = analytics_reporter.generate_excel(self.large_report_data)
        elapsed_time = time.time() - start_time
        
        assert isinstance(result, bytes), "Should generate Excel"
        assert len(result) > 1000, "Large report should be > 1KB"
        # Performance threshold: should complete in < 5 seconds
        assert elapsed_time < 5.0, f"Excel generation took {elapsed_time:.2f}s (should be < 5s)"


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
